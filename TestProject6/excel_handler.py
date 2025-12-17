import config
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import os
from datetime import datetime
import pandas as pd

# Hier habe ich den Dateinamen der Excel-Tabelle festgelegt
FOLDERNAME = config.EXCEL_FOLDER # <--- NEU: Statt "Wetter-Berichte"
FILENAME  = config.EXCEL_FILENAME # <--- NEU: Statt "wetter_report.xlsx"

def save_to_excel(stadt, temp, beschreibung, wind, hum):
    """
    Speichert Wetterdaten in eine Excel-Datei
    """

    # 1. Erst nur den Pfad zum ORDNER bauen
    folder_path = os.path.join(os.getcwd(), FOLDERNAME) 
    
    # 2. WICHTIG!: Prüfe, ob der Ordner existiert, sonst erstellen (mkdir)
    # Ohne das stürzt openpyxl ab, weil es nicht in nicht-existente Ordner speichern kann.
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # 3. Jetzt den Pfad zur Datei im Ordner bauen
    file_path = os.path.join(folder_path, FILENAME) 
    # ### FIX ENDE ###
    
    try:
        # Prüft ob es die Datei schon gibt
        if not os.path.exists(file_path):
            # Wenn keine Datei existiert wird eine neue angelegt
            wb = Workbook()
            # Das Standart leere Blatt entferne ich hier
            ws = wb.active
            wb.title = stadt
        else:
            # Vorhandene Excel-Datei laden
            wb = load_workbook(file_path)
            
        # --- NEUE LOGIK FÜR STÄDTE-TABS ---
        
        # Pfüft ob es für die Stadt schon ein Tabellenblatt gibt
        if stadt in wb.sheetnames:
            # Blatt existiert -> wird es genutzt
            ws = wb[stadt]
        else:
            # Blatt existiert noch nicht -> wir erstellen es
            ws = wb.create_sheet(stadt)
            
            # Überschriften in der Excel definieren
            ws.append(["Zeitstempel", "Stadt", "Temperatur", "Beschreibung", "Windgeschwindkeit", "Luftfeuchtigkeit"])
            
            # Die Überschriften Formatieren (Blau und Fett)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # --- ENDE NEUE LOGIK ---
        
        # Daten für die neue Zeile vorbereiten
        zeit=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        neue_zeile = [zeit, stadt, temp, beschreibung, wind, hum]
        
        # Die neue Zeile hinzufügen
        ws.append(neue_zeile)
        
        # Spaltenbreite für bessere Lesbarkeit anpassen
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20

        
        # Die Änderungen speichern
        wb.save(file_path)
        return "Erfolg: Daten gespeichert"
        
    except PermissionError: # Fehler abfangen falls Datei offen ist
        raise Exception(f"Die Datei '{FILENAME}' ist bereits geöffnet. Bitte schließen Sie die Datei und versuchen Sie es erneut.")
    except Exception as e: # Alle anderen Fehler abfangen
        raise Exception(f"Fehler beim Speichern der Daten: {e}")